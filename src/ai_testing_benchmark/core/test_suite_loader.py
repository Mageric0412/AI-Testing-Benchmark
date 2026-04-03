"""
测试套件加载器 - 支持XLSX多Sheet导入。
"""

from typing import Dict, List, Any, Optional, Tuple
from pathlib import Path
import logging
from dataclasses import dataclass, field
from enum import Enum

try:
    import openpyxl
    from openpyxl.workbook.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Workbook = None
    Worksheet = None

# 默认配置
DEFAULT_MAX_FILE_SIZE_MB = 50

import yaml
import json


class SheetType(Enum):
    """Sheet类型枚举。"""
    TEST_CASES = "test_cases"
    CONFIG = "config"
    SCENARIOS = "scenarios"
    METRICS = "metrics"
    THRESHOLDS = "thresholds"
    CUSTOM = "custom"


@dataclass
class TestCase:
    """测试用例结构。"""
    id: str
    scenario_id: str
    phase: str
    description: str
    input: Dict[str, Any]
    expected_output: Optional[Dict[str, Any]] = None
    metadata: Dict[str, Any] = field(default_factory=dict)
    priority: str = "P1"
    tags: List[str] = field(default_factory=list)
    timeout_ms: int = 30000
    retry_count: int = 0


@dataclass
class Scenario:
    """场景结构。"""
    id: str
    name: str
    description: str
    test_case_ids: List[str]
    metadata: Dict[str, Any] = field(default_factory=dict)


@dataclass
class LoadedTestSuite:
    """加载的测试套件。"""
    name: str
    version: str
    test_cases: List[TestCase]
    scenarios: List[Scenario]
    config: Dict[str, Any]
    raw_data: Dict[str, Any] = field(default_factory=dict)
    metadata: Dict[str, Any] = field(default_factory=dict)


class TestSuiteLoader:
    """
    测试套件加载器。

    支持从XLSX/JSON/YAML文件加载测试套件。
    支持多Sheet导入。
    """

    SHEET_TYPE_MAPPING = {
        "testcases": SheetType.TEST_CASES,
        "test_cases": SheetType.TEST_CASES,
        "tests": SheetType.TEST_CASES,
        "config": SheetType.CONFIG,
        "configuration": SheetType.CONFIG,
        "scenarios": SheetType.SCENARIOS,
        "metrics": SheetType.METRICS,
        "thresholds": SheetType.THRESHOLDS,
    }

    REQUIRED_TEST_CASE_COLUMNS = ["id", "scenario_id", "phase", "description"]
    OPTIONAL_TEST_CASE_COLUMNS = ["input", "expected_output", "metadata", "priority", "tags", "timeout_ms", "retry_count"]

    def __init__(self, config_path: Optional[str] = None):
        """
        初始化测试套件加载器。

        参数:
            config_path: 可选的配置文件路径
        """
        self.logger = logging.getLogger(__name__)
        self.config = self._load_config(config_path) if config_path else {}

    def _load_config(self, config_path: str) -> Dict[str, Any]:
        """加载配置文件。"""
        path = Path(config_path)
        if path.suffix in [".yaml", ".yml"]:
            with open(path, 'r') as f:
                return yaml.safe_load(f) or {}
        elif path.suffix == ".json":
            with open(path, 'r') as f:
                return json.load(f)
        return {}

    def load_from_xlsx(
        self,
        file_path: str,
        sheet_names: Optional[List[str]] = None,
        validate: bool = True
    ) -> LoadedTestSuite:
        """
        从XLSX文件加载测试套件。

        参数:
            file_path: XLSX文件路径
            sheet_names: 要加载的Sheet名称列表，None表示全部
            validate: 是否验证数据

        返回:
            LoadedTestSuite对象

        异常:
            ImportError: 当openpyxl未安装时
            ValueError: 当数据格式无效时
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl未安装。请运行: pip install openpyxl"
            )

        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")

        # 文件大小检查
        file_size_mb = path.stat().st_size / (1024 * 1024)
        if file_size_mb > DEFAULT_MAX_FILE_SIZE_MB:
            raise ValueError(
                f"文件过大 ({file_size_mb:.1f}MB > {DEFAULT_MAX_FILE_SIZE_MB}MB)。"
                f"请使用小于 {DEFAULT_MAX_FILE_SIZE_MB}MB 的文件。"
            )

        # 空文件检查
        if path.stat().st_size == 0:
            raise ValueError("文件为空，请提供有效的XLSX文件。")

        self.logger.info(f"正在加载XLSX文件: {file_path}")

        wb = openpyxl.load_workbook(file_path, data_only=True)

        # 确定要加载的Sheet
        if sheet_names is None:
            sheet_names = wb.sheetnames

        # 按类型解析Sheet
        sheet_data = {}
        for sheet_name in sheet_names:
            if sheet_name in wb.sheetnames:
                sheet_data[sheet_name] = self._parse_sheet(wb[sheet_name])

        # 提取测试套件元数据
        suite_meta = self._extract_suite_metadata(sheet_data)

        # 解析测试用例
        test_cases = []
        scenarios = []

        for sheet_name, data in sheet_data.items():
            sheet_type = self._detect_sheet_type(sheet_name, data)

            if sheet_type == SheetType.TEST_CASES:
                parsed_cases = self._parse_test_cases(data)
                test_cases.extend(parsed_cases)
            elif sheet_type == SheetType.SCENARIOS:
                parsed_scenarios = self._parse_scenarios(data, test_cases)
                scenarios.extend(parsed_scenarios)

        # 提取配置
        config = {}
        if "config" in sheet_data:
            config = sheet_data["config"]
        elif "Config" in sheet_data:
            config = sheet_data["Config"]
        elif "Configuration" in sheet_data:
            config = sheet_data["Configuration"]

        suite = LoadedTestSuite(
            name=suite_meta.get("name", path.stem),
            version=suite_meta.get("version", "1.0.0"),
            test_cases=test_cases,
            scenarios=scenarios,
            config=config,
            raw_data=sheet_data,
            metadata=suite_meta
        )

        if validate:
            self._validate_suite(suite)

        self.logger.info(
            f"加载完成: {len(test_cases)}个测试用例, {len(scenarios)}个场景"
        )

        return suite

    def _parse_sheet(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """解析Sheet为字典列表。"""
        data = []

        # 获取表头
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value)

        # 解析每一行
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            if all(cell.value is None for cell in row):
                continue  # 跳过空行

            row_data = {}
            for header, cell in zip(headers, row):
                if header is not None:
                    value = cell.value
                    # 尝试解析JSON字符串
                    if isinstance(value, str) and value.startswith("{") or isinstance(value, str) and value.startswith("["):
                        try:
                            value = json.loads(value)
                        except json.JSONDecodeError:
                            pass
                    row_data[header] = value
            data.append(row_data)

        return data

    def _detect_sheet_type(self, sheet_name: str, data: List[Dict]) -> SheetType:
        """检测Sheet类型。"""
        # 通过名称检测
        name_lower = sheet_name.lower()
        for key, sheet_type in self.SHEET_TYPE_MAPPING.items():
            if key in name_lower:
                return sheet_type

        # 通过数据结构检测
        if not data:
            return SheetType.CUSTOM

        first_row = data[0]

        # 检查是否包含测试用例必需列
        test_case_cols = set(self.REQUIRED_TEST_CASE_COLUMNS)
        if test_case_cols.issubset(first_row.keys()):
            return SheetType.TEST_CASES

        # 检查是否有特定的场景列
        if "scenario_id" in first_row or "scenario_name" in first_row:
            return SheetType.SCENARIOS

        return SheetType.CUSTOM

    def _parse_test_cases(self, data: List[Dict[str, Any]]) -> List[TestCase]:
        """解析测试用例。"""
        test_cases = []

        for row_idx, row in enumerate(data, start=2):
            try:
                # 检查必需列
                missing_cols = set(self.REQUIRED_TEST_CASE_COLUMNS) - set(row.keys())
                if missing_cols:
                    self.logger.warning(
                        f"行{row_idx}: 缺少必需列 {missing_cols}，跳过"
                    )
                    continue

                # 解析input字段
                input_data = row.get("input", {})
                if isinstance(input_data, str):
                    if input_data.startswith("{"):
                        try:
                            input_data = json.loads(input_data)
                        except json.JSONDecodeError:
                            input_data = {"raw": input_data}
                    else:
                        input_data = {"text": input_data}

                # 解析expected_output字段
                expected_output = row.get("expected_output")
                if isinstance(expected_output, str) and expected_output.startswith("{"):
                    try:
                        expected_output = json.loads(expected_output)
                    except json.JSONDecodeError:
                        expected_output = None

                # 解析metadata字段
                metadata = row.get("metadata", {})
                if isinstance(metadata, str) and metadata.startswith("{"):
                    try:
                        metadata = json.loads(metadata)
                    except json.JSONDecodeError:
                        metadata = {}

                # 解析tags
                tags = row.get("tags", [])
                if isinstance(tags, str):
                    tags = [t.strip() for t in tags.split(",")]

                test_case = TestCase(
                    id=str(row["id"]),
                    scenario_id=str(row["scenario_id"]),
                    phase=str(row["phase"]),
                    description=str(row["description"]),
                    input=input_data or {},
                    expected_output=expected_output,
                    metadata=metadata or {},
                    priority=str(row.get("priority", "P1")),
                    tags=tags or [],
                    timeout_ms=int(row.get("timeout_ms", 30000)),
                    retry_count=int(row.get("retry_count", 0))
                )

                test_cases.append(test_case)

            except Exception as e:
                self.logger.warning(f"行{row_idx}: 解析失败 - {str(e)}")
                continue

        return test_cases

    def _parse_scenarios(
        self,
        data: List[Dict[str, Any]],
        test_cases: List[TestCase]
    ) -> List[Scenario]:
        """解析场景。"""
        scenarios = []
        test_case_ids = {tc.id for tc in test_cases}

        for row_idx, row in enumerate(data, start=2):
            try:
                scenario_id = str(row["id"])
                test_case_ids_str = row.get("test_case_ids", "")

                if isinstance(test_case_ids_str, str):
                    scenario_tc_ids = [
                        tid.strip() for tid in test_case_ids_str.split(",")
                        if tid.strip()
                    ]
                elif isinstance(test_case_ids_str, list):
                    scenario_tc_ids = test_case_ids_str
                else:
                    scenario_tc_ids = []

                # 验证测试用例ID存在
                invalid_ids = set(scenario_tc_ids) - test_case_ids
                if invalid_ids:
                    self.logger.warning(
                        f"场景{scenario_id}: 引用了不存在的测试用例 {invalid_ids}"
                    )

                scenario = Scenario(
                    id=scenario_id,
                    name=str(row.get("name", scenario_id)),
                    description=str(row.get("description", "")),
                    test_case_ids=[tid for tid in scenario_tc_ids if tid in test_case_ids],
                    metadata=row.get("metadata", {})
                )

                scenarios.append(scenario)

            except Exception as e:
                self.logger.warning(f"场景行{row_idx}: 解析失败 - {str(e)}")
                continue

        return scenarios

    def _extract_suite_metadata(self, sheet_data: Dict[str, Any]) -> Dict[str, Any]:
        """提取测试套件元数据。"""
        meta = {
            "name": "Unnamed Test Suite",
            "version": "1.0.0"
        }

        # 从Config Sheet提取
        config_sheets = ["config", "Config", "Configuration"]
        for sheet_name in config_sheets:
            if sheet_name in sheet_data and sheet_data[sheet_name]:
                first_row = sheet_data[sheet_name][0]
                if "name" in first_row:
                    meta["name"] = first_row["name"]
                if "version" in first_row:
                    meta["version"] = first_row["version"]

        # 从test_cases Sheet的第一行提取
        for sheet_name in ["testcases", "test_cases", "tests"]:
            if sheet_name in sheet_data and sheet_data[sheet_name]:
                first_row = sheet_data[sheet_name][0]
                for key in ["suite_name", "suite_name", "benchmark_name"]:
                    if key in first_row:
                        meta["name"] = first_row[key]

        return meta

    def _validate_suite(self, suite: LoadedTestSuite) -> None:
        """验证测试套件的完整性。"""
        # 检查是否有测试用例
        if not suite.test_cases:
            raise ValueError("测试套件不包含任何测试用例")

        # 检查是否有重复的测试用例ID
        ids = [tc.id for tc in suite.test_cases]
        duplicates = set(ids) - set(ids)
        if duplicates:
            raise ValueError(f"存在重复的测试用例ID: {duplicates}")

        # 检查场景引用的测试用例是否存在
        scenario_tc_ids = set()
        for scenario in suite.scenarios:
            scenario_tc_ids.update(scenario.test_case_ids)

        missing_tc_ids = scenario_tc_ids - set(ids)
        if missing_tc_ids:
            self.logger.warning(
                f"场景引用了不存在的测试用例: {missing_tc_ids}"
            )

    def save_to_xlsx(
        self,
        suite: LoadedTestSuite,
        output_path: str,
        include_config_sheet: bool = True
    ) -> None:
        """
        将测试套件保存为XLSX文件。

        参数:
            suite: 测试套件
            output_path: 输出文件路径
            include_config_sheet: 是否包含配置Sheet
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl未安装")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # 移除默认sheet

        # 保存测试用例Sheet
        ws_tests = wb.create_sheet("test_cases")
        headers = self.REQUIRED_TEST_CASE_COLUMNS + self.OPTIONAL_TEST_CASE_COLUMNS
        ws_tests.append(headers)

        for tc in suite.test_cases:
            row = [
                tc.id,
                tc.scenario_id,
                tc.phase,
                tc.description,
                json.dumps(tc.input) if tc.input else "",
                json.dumps(tc.expected_output) if tc.expected_output else "",
                json.dumps(tc.metadata),
                tc.priority,
                ",".join(tc.tags),
                tc.timeout_ms,
                tc.retry_count
            ]
            ws_tests.append(row)

        # 保存场景Sheet
        if suite.scenarios:
            ws_scenarios = wb.create_sheet("scenarios")
            ws_scenarios.append(["id", "name", "description", "test_case_ids", "metadata"])
            for s in suite.scenarios:
                ws_scenarios.append([
                    s.id,
                    s.name,
                    s.description,
                    ",".join(s.test_case_ids),
                    json.dumps(s.metadata)
                ])

        # 保存配置Sheet
        if include_config_sheet and suite.config:
            ws_config = wb.create_sheet("config")
            ws_config.append(["key", "value"])
            for key, value in suite.config.items():
                ws_config.append([key, json.dumps(value) if isinstance(value, dict) else value])

        wb.save(output_path)
        self.logger.info(f"测试套件已保存至: {output_path}")

    def get_sheet_info(self, file_path: str) -> List[Dict[str, Any]]:
        """
        获取XLSX文件的Sheet信息。

        参数:
            file_path: XLSX文件路径

        返回:
            Sheet信息列表
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl未安装")

        wb = openpyxl.load_workbook(file_path, data_only=True)

        sheet_info = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            info = {
                "name": sheet_name,
                "type": self._detect_sheet_type(sheet_name, self._parse_sheet(sheet)).value,
                "rows": sheet.max_row,
                "columns": sheet.max_column,
                "headers": [cell.value for cell in sheet[1]] if sheet.max_row > 0 else []
            }
            sheet_info.append(info)

        return sheet_info


def load_test_suite(
    file_path: str,
    config_path: Optional[str] = None,
    sheet_names: Optional[List[str]] = None
) -> LoadedTestSuite:
    """
    便捷函数：加载测试套件。

    自动检测文件类型并加载。

    参数:
        file_path: 文件路径
        config_path: 可选的配置文件路径
        sheet_names: 要加载的Sheet名称

    返回:
        LoadedTestSuite对象
    """
    loader = TestSuiteLoader(config_path)
    path = Path(file_path)

    if path.suffix.lower() in [".xlsx", ".xls"]:
        return loader.load_from_xlsx(file_path, sheet_names)
    elif path.suffix in [".yaml", ".yml"]:
        return loader._load_from_yaml(path)
    elif path.suffix == ".json":
        return loader._load_from_json(path)
    else:
        raise ValueError(f"不支持的文件格式: {path.suffix}")
